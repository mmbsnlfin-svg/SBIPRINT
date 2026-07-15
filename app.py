import io
import os
import re
import shutil
import subprocess
import tempfile
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_OUTPUT_HEADERS = [
    "Sr. No.", "ACCOUNT _NUM", "LC ID", "Bill From", "Bill To", "Days",
    "Branch Code", "Branch Name", "C Type", "Port BW",
    "Annual Recurring Charges", "Quarterly Charges", "NTU Chg /Modem Chg",
    "NOFN charges", "IDR / Submarine Charges",
    "Total Quarterly charges Gross", "GST 18%", "Net Payable After Tax",
    "GST STATE", "Parent BA", "PO NO", "PO Date",
]

HEADER_ALIASES = {
    "ACCOUNT _NUM": {"ACCOUNTNUM", "ACCOUNTNUMBER", "ACCOUNTNO"},
    "LC ID": {"LCID"},
    "Bill From": {"BILLFROM"},
    "Bill To": {"BILLTO"},
    "Days": {"DAYS"},
    "Branch Code": {"BRANCHCODE"},
    "Branch Name": {"BRANCHNAME"},
    "C Type": {"CTYPE"},
    "Port BW": {"PORTBW", "PORTBANDWIDTH"},
    "Annual Recurring Charges": {
        "ANNUALRECURRINGCHARGES", "ANNUALRECURRINGCHARGESAFTERDISCOUNT", "ARC"
    },
    "Quarterly Charges": {"QUARTERLYCHARGES"},
    "NTU Chg /Modem Chg": {
        "NTUCHGMODEMCHG", "NTUCHARGEMODEMCHARGE", "NTUCHARGESMODEMCHARGES"
    },
    "NOFN charges": {"NOFNCHARGES"},
    "IDR / Submarine Charges": {
        "IDRSUBMARINECHARGES", "IDRCHARGES", "SUBMARINECHARGES"
    },
    "Total Quarterly charges Gross": {
        "TOTALQUARTERLYCHARGESGROSS", "TOTALQUARTERLYCHARGES"
    },
    "GST 18%": {"GST18", "GST18PERCENT"},
    "Net Payable After Tax": {"NETPAYABLEAFTERTAX"},
    "GST STATE": {"GSTSTATE", "GSTSTATECODE", "STATE"},
    "Parent BA": {"PARENTBA"},
    "PO NO": {"PONO", "PONUMBER"},
    "PO Date": {"PODATE"},
}

SUMMARY_HEADERS = [
    "S L NO", "State Code", "Parent", "Invoice no", "Invoice date",
    "BSNL GST No.", "SBI GST No.", "Ckts", "Amount", "Gross Total",
    "CGST @ 9%", "SGST @ 9%", "Total GST @ 18%",
    "Total Amount including GST",
]

BLACK = "000000"
GREEN = "92D050"
HEADER_FILL = PatternFill("solid", fgColor=GREEN)
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN_SIDE = Side(style="thin", color=BLACK)
MEDIUM_SIDE = Side(style="medium", color=BLACK)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
MEDIUM_BORDER = Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE, bottom=MEDIUM_SIDE)

COLUMN_WIDTHS = {
    "Sr. No.": 7, "ACCOUNT _NUM": 15, "LC ID": 14, "Bill From": 13,
    "Bill To": 13, "Days": 7, "Branch Code": 11, "Branch Name": 29,
    "C Type": 9, "Port BW": 11, "Annual Recurring Charges": 18,
    "Quarterly Charges": 17, "NTU Chg /Modem Chg": 16,
    "NOFN charges": 14, "IDR / Submarine Charges": 17,
    "Total Quarterly charges Gross": 20, "GST 18%": 15,
    "Net Payable After Tax": 19, "GST STATE": 11, "Parent BA": 14,
    "PO NO": 14, "PO Date": 13,
}

FORMULA_HEADERS = {
    "Days", "Quarterly Charges", "Total Quarterly charges Gross",
    "GST 18%", "Net Payable After Tax",
}

TOTAL_HEADERS = {
    "Annual Recurring Charges", "Quarterly Charges", "NTU Chg /Modem Chg",
    "NOFN charges", "IDR / Submarine Charges",
    "Total Quarterly charges Gross", "GST 18%", "Net Payable After Tax",
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").replace("\u2007", " ").replace("\u202f", " ")
    return re.sub(r"[^A-Z0-9]", "", text.upper())


def clean_filename(value: Any) -> str:
    text = re.sub(r'[\\/*?:"<>|]', "", str(value).strip())
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text[:100] or "Blank_Value"


def unique_filename(base_name: str, used: set[str], extension: str) -> str:
    candidate = f"{base_name}{extension}"
    count = 2
    while candidate.lower() in used:
        candidate = f"{base_name}_{count}{extension}"
        count += 1
    used.add(candidate.lower())
    return candidate


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)


def find_header_row(ws, max_scan_rows: int = 10) -> int:
    """Locate the table header efficiently in read-only mode."""
    required_aliases = set().union(*HEADER_ALIASES.values())
    best_row, best_count = 0, 0
    for row_no, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows), values_only=True),
        start=1,
    ):
        found = sum(1 for value in row if normalize_header(value) in required_aliases)
        if found > best_count:
            best_row, best_count = row_no, found
    if best_count < 5:
        raise ValueError("Could not identify the Excel header row.")
    return best_row


def build_source_header_map(ws, header_row: int) -> dict[str, int]:
    header_values = next(
        ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )
    normalized_to_col: dict[str, int] = {}
    for col, value in enumerate(header_values, start=1):
        normalized = normalize_header(value)
        if normalized and normalized not in normalized_to_col:
            normalized_to_col[normalized] = col

    mapping, missing = {}, []
    for display_header in REQUIRED_OUTPUT_HEADERS[1:]:
        matched = next(
            (normalized_to_col[a] for a in HEADER_ALIASES[display_header] if a in normalized_to_col),
            None,
        )
        if matched is None:
            missing.append(display_header)
        else:
            mapping[display_header] = matched
    if missing:
        raise ValueError("Required columns not found: " + ", ".join(missing))
    return mapping


def output_column_map() -> dict[str, int]:
    return {header: i for i, header in enumerate(REQUIRED_OUTPUT_HEADERS, start=1)}


def excel_ref(header: str, row: int, out_map: dict[str, int]) -> str:
    return f"{get_column_letter(out_map[header])}{row}"


def data_formula(header: str, row: int, out_map: dict[str, int]) -> str:
    if header == "Days":
        return f"={excel_ref('Bill To', row, out_map)}-{excel_ref('Bill From', row, out_map)}+1"
    if header == "Quarterly Charges":
        return f"={excel_ref('Annual Recurring Charges', row, out_map)}/4"
    if header == "Total Quarterly charges Gross":
        start = excel_ref("Quarterly Charges", row, out_map)
        end = excel_ref("IDR / Submarine Charges", row, out_map)
        return f"=SUM({start}:{end})"
    if header == "GST 18%":
        return f"={excel_ref('Total Quarterly charges Gross', row, out_map)}*18%"
    if header == "Net Payable After Tax":
        return f"={excel_ref('Total Quarterly charges Gross', row, out_map)}+{excel_ref('GST 18%', row, out_map)}"
    raise KeyError(header)


def to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "--"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def title_text(source_title: Any, split_value: Any) -> str:
    text = str(source_title or "").strip()
    state = str(split_value).strip()
    if state and not re.search(rf"(?:-|\s){re.escape(state)}\s*$", text, flags=re.IGNORECASE):
        text = f"{text} - {state}" if text else f"GST STATE: {state}"
    return text


def style_output_sheet(ws, total_row: int, body_font_size: int = 12, bottom_margin: float = 1.25) -> None:
    max_col = len(REQUIRED_OUTPUT_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title = ws.cell(1, 1)
    title.font = Font(name="Calibri", size=20, bold=True, color=BLACK)
    title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title.border = MEDIUM_BORDER
    ws.row_dimensions[1].height = 48

    for col, header in enumerate(REQUIRED_OUTPUT_HEADERS, start=1):
        cell = ws.cell(2, col)
        cell.value = header
        cell.font = Font(name="Calibri", size=12, bold=True, color=BLACK)
        cell.fill = HEADER_FILL
        cell.border = MEDIUM_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[header]
    ws.row_dimensions[2].height = 72

    # Apply styles by row/column without copying source styles. This is much faster.
    body_font = Font(name="Calibri", size=body_font_size, color=BLACK)
    total_font = Font(name="Calibri", size=body_font_size, bold=True, color=BLACK)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row in range(3, total_row + 1):
        is_total = row == total_row
        for col in range(1, max_col + 1):
            header = REQUIRED_OUTPUT_HEADERS[col - 1]
            cell = ws.cell(row, col)
            cell.font = total_font if is_total else body_font
            cell.border = MEDIUM_BORDER if is_total else THIN_BORDER
            cell.alignment = left if header == "Branch Name" else center
        ws.row_dimensions[row].height = 40 if is_total else 36

    out_map = output_column_map()
    for header in ("Bill From", "Bill To", "PO Date"):
        col = out_map[header]
        for row in range(3, total_row):
            ws.cell(row, col).number_format = "dd-mmm-yyyy"

    for header in TOTAL_HEADERS:
        col = out_map[header]
        for row in range(3, total_row + 1):
            ws.cell(row, col).number_format = '#,##0.00;[Red]-#,##0.00;"-"'

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{total_row - 1}"
    ws.print_title_rows = "1:2"
    ws.print_area = f"A1:{get_column_letter(max_col)}{total_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.scale = None
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = bottom_margin
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.15
    ws.oddFooter.center.text = "Page &P of &N"
    ws.evenFooter.center.text = "Page &P of &N"
    ws.sheet_view.showGridLines = False


def read_summary_master(master_bytes: bytes | None) -> dict[tuple[str, str], dict[str, Any]]:
    if not master_bytes:
        return {}
    wb = load_workbook(io.BytesIO(master_bytes), data_only=True, read_only=True)
    try:
        ws = wb.active
        header_row = None
        col_map: dict[str, int] = {}
        aliases = {
            "STATE": {"STATECODE", "GSTSTATE", "STATE"},
            "PARENT": {"PARENT", "PARENTBA"},
            "BSNLGST": {"BSNLGSTNO", "BSNLGSTIN", "BSNLGSTNUMBER"},
            "SBIGST": {"SBIGSTNO", "SBIGSTIN", "SBIGSTNUMBER"},
        }
        for r, values in enumerate(
            ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True), start=1
        ):
            row_values = {normalize_header(v): c for c, v in enumerate(values, start=1)}
            found: dict[str, int] = {}
            for key, names in aliases.items():
                col = next((row_values[n] for n in names if n in row_values), None)
                if col:
                    found[key] = col
            if "STATE" in found:
                header_row, col_map = r, found
                break
        if not header_row:
            return {}

        result: dict[tuple[str, str], dict[str, Any]] = {}
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            state = values[col_map["STATE"] - 1] if len(values) >= col_map["STATE"] else None
            state_text = str(state).strip() if state is not None else ""
            if not state_text or state_text.lower() == "total":
                continue
            parent = values[col_map["PARENT"] - 1] if "PARENT" in col_map and len(values) >= col_map["PARENT"] else None
            parent_text = str(parent).strip() if parent is not None else ""
            gst_values = {
                "BSNL GST No.": values[col_map["BSNLGST"] - 1] if "BSNLGST" in col_map and len(values) >= col_map["BSNLGST"] else None,
                "SBI GST No.": values[col_map["SBIGST"] - 1] if "SBIGST" in col_map and len(values) >= col_map["SBIGST"] else None,
            }
            result[(state_text, parent_text)] = gst_values
            result.setdefault((state_text, ""), gst_values)
        return result
    finally:
        wb.close()


def build_summary_workbook(summary_rows: list[dict[str, Any]], summary_title: str, body_font_size: int = 12, bottom_margin: float = 1.25) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    max_col = len(SUMMARY_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1).value = summary_title
    ws.cell(1, 1).font = Font(name="Calibri", size=20, bold=True, color=BLACK)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(1, 1).border = MEDIUM_BORDER
    ws.row_dimensions[1].height = 48

    for col, header in enumerate(SUMMARY_HEADERS, start=1):
        c = ws.cell(2, col, header)
        c.font = Font(name="Calibri", size=12, bold=True, color=BLACK)
        c.fill = HEADER_FILL
        c.border = MEDIUM_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 58

    widths = [8, 11, 14, 20, 14, 20, 20, 10, 16, 16, 15, 15, 17, 20]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    row_no = 3
    for item in summary_rows:
        values = [
            item["S L NO"], item["State Code"], item.get("Parent"), item.get("Invoice no"),
            item.get("Invoice date"), item.get("BSNL GST No."), item.get("SBI GST No."),
            item["Ckts"], item["Amount"], item["Gross Total"], item["CGST @ 9%"],
            item["SGST @ 9%"], item["Total GST @ 18%"], item["Total Amount including GST"],
        ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row_no, col, value)
            c.font = Font(name="Calibri", size=body_font_size, color=BLACK)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row_no].height = 36
        row_no += 1

    total_row = row_no
    ws.cell(total_row, 7).value = "Total"
    for col in range(8, 15):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = f"=SUM({letter}3:{letter}{total_row - 1})"
    for col in range(1, max_col + 1):
        c = ws.cell(total_row, col)
        c.font = Font(name="Calibri", size=body_font_size, bold=True, color=BLACK)
        c.border = MEDIUM_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[total_row].height = 40

    for row in range(3, total_row + 1):
        ws.cell(row, 5).number_format = "dd-mmm-yyyy"
        for col in range(9, 15):
            ws.cell(row, col).number_format = '#,##0.00;[Red]-#,##0.00;"-"'

    ws.freeze_panes = "A3"
    ws.print_title_rows = "1:2"
    ws.print_area = f"A1:N{total_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = bottom_margin
    ws.oddFooter.center.text = "Page &P of &N"
    ws.sheet_view.showGridLines = False

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def _value(values: tuple, col: int) -> Any:
    return values[col - 1] if col and len(values) >= col else None


def split_excel_file(
    uploaded_bytes: bytes,
    original_name: str,
    summary_master_bytes: bytes | None,
    summary_title_override: str,
    progress_callback=None,
    body_font_size: int = 12,
    bottom_margin: float = 1.25,
) -> tuple[dict[str, bytes], list[str]]:
    """Optimized for 20,000+ rows.

    The input workbook is read once in read-only mode. Rows are grouped by state
    during that one pass, avoiding the previous states × rows nested scan.
    """
    source_wb = load_workbook(io.BytesIO(uploaded_bytes), data_only=True, read_only=True)
    warnings: list[str] = []
    try:
        source_ws = source_wb.active
        header_row = find_header_row(source_ws)
        source_map = build_source_header_map(source_ws, header_row)
        out_map = output_column_map()
        split_source_col = source_map["GST STATE"]

        title_row = max(1, header_row - 1)
        title_values = next(source_ws.iter_rows(min_row=title_row, max_row=title_row, values_only=True))
        source_title = title_values[0] if title_values else ""

        # One pass through the input. Keep only required values, not styles/unused columns.
        state_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
        summary_groups: dict[tuple[str, str], dict[str, Any]] = {}
        total_rows = max(1, source_ws.max_row - header_row)

        for processed, values in enumerate(
            source_ws.iter_rows(min_row=header_row + 1, values_only=True), start=1
        ):
            state_raw = _value(values, split_source_col)
            state_text = str(state_raw).strip() if state_raw is not None else ""
            if not state_text or state_text.lower() in {"none", "nan", "total"}:
                continue

            lc_value = _value(values, source_map["LC ID"])
            # Skip blank rows and source total rows.
            if lc_value is None or str(lc_value).strip().lower() in {"", "total"}:
                continue

            row_data = {
                header: _value(values, source_map[header])
                for header in REQUIRED_OUTPUT_HEADERS[1:]
                if header not in FORMULA_HEADERS
            }
            state_rows[state_text].append(row_data)

            parent_raw = row_data.get("Parent BA")
            parent_text = str(parent_raw).strip() if parent_raw is not None else ""
            group = summary_groups.setdefault((state_text, parent_text), {
                "State Code": state_text, "Parent": parent_raw, "Ckts": 0,
                "Annual": 0.0, "NTU": 0.0, "NOFN": 0.0, "IDR": 0.0,
            })
            group["Ckts"] += 1
            group["Annual"] += to_number(row_data.get("Annual Recurring Charges"))
            group["NTU"] += to_number(row_data.get("NTU Chg /Modem Chg"))
            group["NOFN"] += to_number(row_data.get("NOFN charges"))
            group["IDR"] += to_number(row_data.get("IDR / Submarine Charges"))

            if progress_callback and (processed % 500 == 0 or processed == total_rows):
                progress_callback(min(processed / total_rows, 0.70), f"Reading row {processed:,} of about {total_rows:,}")

        if not state_rows:
            raise ValueError("No valid GST STATE / LC ID data rows were found.")

        source_wb.close()
        summary_master = read_summary_master(summary_master_bytes)
        output_files: dict[str, bytes] = {}
        used_names: set[str] = set()
        states = sorted(state_rows)

        for state_index, state_text in enumerate(states, start=1):
            rows = state_rows[state_text]
            new_wb = Workbook()
            ws = new_wb.active
            ws.title = clean_filename(state_text)[:31]
            ws.cell(1, 1).value = title_text(source_title, state_text)
            for col, header in enumerate(REQUIRED_OUTPUT_HEADERS, start=1):
                ws.cell(2, col).value = header

            for serial, row_data in enumerate(rows, start=1):
                destination_row = serial + 2
                ws.cell(destination_row, out_map["Sr. No."]).value = serial
                for header in REQUIRED_OUTPUT_HEADERS[1:]:
                    dst = ws.cell(destination_row, out_map[header])
                    if header in FORMULA_HEADERS:
                        dst.value = data_formula(header, destination_row, out_map)
                    else:
                        dst.value = row_data.get(header)

            total_row = len(rows) + 3
            ws.cell(total_row, out_map["Port BW"]).value = "Total"
            for header in TOTAL_HEADERS:
                col = out_map[header]
                letter = get_column_letter(col)
                ws.cell(total_row, col).value = f"=SUM({letter}3:{letter}{total_row - 1})"
            style_output_sheet(ws, total_row, body_font_size=body_font_size, bottom_margin=bottom_margin)

            out = io.BytesIO()
            new_wb.save(out)
            new_wb.close()
            filename = unique_filename(clean_filename(state_text), used_names, ".xlsx")
            output_files[filename] = out.getvalue()

            if progress_callback:
                progress_callback(0.70 + 0.25 * state_index / len(states), f"Created {filename} ({len(rows):,} rows)")

        summary_rows: list[dict[str, Any]] = []
        for serial_no, ((state_text, parent_text), group) in enumerate(
            sorted(summary_groups.items(), key=lambda item: (item[0][0], item[0][1])), start=1
        ):
            quarterly = group["Annual"] / 4
            gross = quarterly + group["NTU"] + group["NOFN"] + group["IDR"]
            gst = gross * 0.18
            meta = summary_master.get((state_text, parent_text)) or summary_master.get((state_text, ""), {})
            summary_rows.append({
                "S L NO": serial_no, "State Code": state_text,
                "Parent": group.get("Parent") or parent_text,
                "Invoice no": None, "Invoice date": None,
                "BSNL GST No.": meta.get("BSNL GST No."),
                "SBI GST No.": meta.get("SBI GST No."),
                "Ckts": group["Ckts"], "Amount": gross, "Gross Total": gross,
                "CGST @ 9%": gross * 0.09, "SGST @ 9%": gross * 0.09,
                "Total GST @ 18%": gst, "Total Amount including GST": gross + gst,
            })

        summary_title = summary_title_override.strip() or f"Summary - {str(source_title or '').strip()}"
        output_files["Summary.xlsx"] = build_summary_workbook(summary_rows, summary_title, body_font_size=body_font_size, bottom_margin=bottom_margin)
        if progress_callback:
            progress_callback(1.0, "Excel generation completed")
        return output_files, warnings
    finally:
        try:
            source_wb.close()
        except Exception:
            pass

def find_libreoffice() -> str | None:
    for command in ("libreoffice", "soffice"):
        found = shutil.which(command)
        if found:
            return found
    for path in (
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ):
        if os.path.exists(path):
            return path
    return None


def convert_excels_to_pdfs(excel_files: dict[str, bytes]) -> tuple[dict[str, bytes], list[str]]:
    soffice = find_libreoffice()
    if not soffice:
        raise RuntimeError("LibreOffice is not installed. Add 'libreoffice-calc' to packages.txt.")
    pdf_files, errors = {}, []
    with tempfile.TemporaryDirectory(prefix="excel_pdf_printable_") as temp_dir:
        root = Path(temp_dir)
        excel_dir, pdf_dir = root / "excel", root / "pdf"
        excel_dir.mkdir(); pdf_dir.mkdir()
        for index, (filename, data) in enumerate(excel_files.items(), start=1):
            excel_path = excel_dir / filename
            excel_path.write_bytes(data)
            profile_dir = root / f"profile_{index}_{uuid.uuid4().hex}"
            profile_dir.mkdir()
            try:
                command = [
                    soffice, "--headless", "--nologo", "--nofirststartwizard", "--norestore",
                    f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
                    "--convert-to", "pdf:calc_pdf_Export", "--outdir", str(pdf_dir), str(excel_path),
                ]
                result = subprocess.run(command, capture_output=True, text=True, timeout=240, check=False)
                expected = pdf_dir / f"{excel_path.stem}.pdf"
                if result.returncode != 0 or not expected.exists():
                    raise RuntimeError((result.stderr or result.stdout or "Unknown LibreOffice error").strip())
                pdf_files[expected.name] = expected.read_bytes()
            except Exception as exc:
                errors.append(f"{filename}: {exc}")
    return pdf_files, errors


def make_zip(files: dict[str, bytes]) -> bytes:
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        for filename, data in files.items():
            archive.writestr(filename, data)
    return stream.getvalue()


def make_combined_zip(excels: dict[str, bytes], pdfs: dict[str, bytes]) -> bytes:
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        for filename, data in excels.items():
            archive.writestr(f"Excel/{filename}", data)
        for filename, data in pdfs.items():
            archive.writestr(f"PDF/{filename}", data)
    return stream.getvalue()


def initialize_state() -> None:
    for key, value in {"excel_files": {}, "pdf_files": {}, "conversion_errors": []}.items():
        if key not in st.session_state:
            st.session_state[key] = value


def main() -> None:
    st.set_page_config(page_title="Excel Splitter & PDF Generator", page_icon="📊", layout="wide")
    initialize_state()
    st.title("Excel Splitter, State-Parent Summary & PDF Generator")
    st.caption(
        "Creates state-wise Excel/PDF files plus Summary.xlsx. "
        "Summary is grouped by GST State Code and Parent BA. Summary PDF is not generated."
    )

    with st.expander("Output structure", expanded=False):
        st.write(REQUIRED_OUTPUT_HEADERS)
        st.info("Excluded: PO, Type, CGST 9%, SGST 9%, Remarks, Loopback and WAN IP.")

    uploaded_file = st.file_uploader("Upload original Excel file", type=["xlsx", "xlsm"])
    summary_master = st.file_uploader(
        "Upload summary master for BSNL GST No. and SBI GST No. (mapped by State Code and Parent)",
        type=["xlsx", "xlsm"], key="summary_master",
    )
    summary_title = st.text_input("Summary title (optional)")
    p1, p2 = st.columns(2)
    body_font_size = p1.number_input("Data font size", min_value=10, max_value=16, value=12, step=1)
    bottom_margin = p2.number_input(
        "Bottom margin for stamp/sign (inches)", min_value=0.25, max_value=3.0,
        value=1.25, step=0.25, help="Keeps blank space at the bottom of every printed page."
    )
    generate_pdf = st.checkbox("Also generate PDF files", value=False, help="For 20,000+ rows, first generate Excel files. PDF conversion may take several minutes per state.")

    if st.button("Generate Files", type="primary", use_container_width=True):
        if uploaded_file is None:
            st.error("Please upload the original Excel file.")
        else:
            try:
                with st.status("Processing workbook...", expanded=True) as status:
                    st.write("Reading the workbook once and creating state-wise files...")
                    progress_bar = st.progress(0, text="Starting...")
                    def update_progress(value: float, text: str) -> None:
                        progress_bar.progress(int(max(0, min(value, 1)) * 100), text=text)

                    excel_files, warnings = split_excel_file(
                        uploaded_file.getvalue(), uploaded_file.name,
                        summary_master.getvalue() if summary_master else None,
                        summary_title,
                        progress_callback=update_progress,
                        body_font_size=int(body_font_size),
                        bottom_margin=float(bottom_margin),
                    )
                    st.session_state.excel_files = excel_files
                    st.session_state.pdf_files = {}
                    st.session_state.conversion_errors = []
                    st.write(f"Created {len(excel_files)} Excel file(s), including Summary.xlsx.")

                    if generate_pdf:
                        st.write("Converting state-wise Excel files to PDF (Summary.xlsx is excluded)...")
                        state_excel_files = {
                            name: data for name, data in excel_files.items()
                            if name.lower() != "summary.xlsx"
                        }
                        pdf_files, errors = convert_excels_to_pdfs(state_excel_files)
                        st.session_state.pdf_files = pdf_files
                        st.session_state.conversion_errors = errors
                        st.write(f"Created {len(pdf_files)} state-wise PDF file(s). Summary PDF is not generated.")

                    for warning in warnings:
                        st.warning(warning)
                    status.update(label="Processing completed", state="complete")
            except Exception as exc:
                st.session_state.excel_files = {}
                st.session_state.pdf_files = {}
                st.session_state.conversion_errors = []
                st.error(str(exc))

    excel_files = st.session_state.excel_files
    pdf_files = st.session_state.pdf_files
    errors = st.session_state.conversion_errors

    if excel_files:
        st.divider()
        st.subheader("Download Results")
        c1, c2, c3 = st.columns(3)
        c1.metric("Excel files", len(excel_files))
        c2.metric("PDF files", len(pdf_files))
        c3.metric("PDF errors", len(errors))

        if "Summary.xlsx" in excel_files:
            st.download_button(
                "Download Summary Excel", excel_files["Summary.xlsx"], "Summary.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        d1, d2, d3 = st.columns(3)
        d1.download_button("Download Excel ZIP", make_zip(excel_files), "Excel_Files.zip", "application/zip", use_container_width=True)
        if pdf_files:
            d2.download_button("Download PDF ZIP", make_zip(pdf_files), "PDF_Files.zip", "application/zip", use_container_width=True)
            d3.download_button("Download Excel + PDF ZIP", make_combined_zip(excel_files, pdf_files), "Excel_and_PDF_Files.zip", "application/zip", use_container_width=True)

        if errors:
            st.warning("Some PDF files could not be generated.")
            st.code("\n".join(errors), language=None)

    st.divider()
    st.caption("Created by  @ HRUSHIKESH KESALE- ACCOUNTS OFFICER ")


if __name__ == "__main__":
    main()
